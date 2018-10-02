#-------------------------------------------------------------------------------
# Name:        bat_map_cli
# Purpose:
#
# Author:      William Trodd
#
# Created:     09/07/2018
# Copyright:   (c) willi_000 2018
# Licence:     <your licence>
#-------------------------------------------------------------------------------

import openpyxl

def get_sheet():
    path = input("Please enter the workbook file path:")
    wb = openpyxl.load_workbook(path, data_only = True)
    return wb

def display_sheets(workbook):
    counter = 1
    sheet_list = []
    for sheet in workbook:
        sheet_list.append(sheet.title)
        print(counter, sheet_list[counter-1])
        counter += 1
    return sheet_list


def load_sheet(workbook, sheet):
    return(workbook[sheet])

def show_columns(sheet):
    current_cell = ['A','1']
    cell_string = current_cell[0]+current_cell[1]

    column_headers = []
    counter = 1
    while sheet[cell_string].value != None:

        column_headers.append([counter-1,sheet[cell_string].value])
        print(counter, column_headers[counter-1][1])

        current_cell[0] = chr(ord(current_cell[0])+1)

        cell_string = current_cell[0]+current_cell[1]
        counter += 1

    return(column_headers)


def get_column_items(sheet, column_num):
    column_items = []
    for row in sheet.iter_rows(min_row = 2):
        item = row[column_num].value
        if item != None and item != 0:
            if column_items == []:
                column_items.append(item)
            elif item not in column_items:
                column_items.append(item)
    return column_items


def mk_kml(sheet,
           species_cols,
           site_col,
           time_col,
           transect_col,
           image_col,
           lat_col,
           long_col,
           species_colours
           ):

    template = kml_format()
    current_site = ""
    current_transect = 0
    scale = 2.5 ## arbitrary but looks quite good.

    for row in sheet.iter_rows(min_row = 2,
                              max_col = max(species_cols[-1],
                                            site_col,
                                            time_col,
                                            transect_col,
                                            image_col,
                                            lat_col,
                                            long_col
                                            )+1
                              ):

        ##current_site = row[site_col].value
        current_time = row[time_col].value
        ##current_transect = row[transect_col].value
        ##current_img = row[img_col].value
        current_lat = row[lat_col].value
        current_long = row[long_col].value
        current_sp = row[species_cols[0]-1].value


        if current_site != None:
            if current_transect != row[transect_col].value:
                try:
                    output.write('\n    </Folder>\n  </Document>\n</kml>')
                except:
                    print("whoops")
                current_site = row[site_col].value
                current_transect = row[transect_col].value
                marker_num = 0
                output = open(current_site +
                              "_" +
                              str(current_transect) +
                              ".kml","w"
                              )

                output.write('<?xml version="1.0" encoding="UTF-8"?>\n')
                output.write('<kml xmnls="http://www.opengis.net/kml/2.2">\n')
                output.write('  <Document>\n    <Folder>\n')
                output.write('        <name>' +
                             current_site +
                             '</name>\n')
                output.write('            <description>Recordings of species at ' +
                             current_site +
                             '</description>)')

            if current_sp != "NONE":
                print("adding species record")
                marker_num += 1
                output.write(template.format(current_site,
                                             scale,
                                             species_colours[current_sp],
                                             marker_num,
                                             current_time,
                                             current_lat,
                                             current_long)   ## writing wrong details to kml but colours work!!!!
                            )
    try:
        output.write('\n    </Folder>\n  </Document>\n</kml>')
        output.close()
        print("closing")
    except:
        print("nothing to close")






def kml_format():

    file_in = open("kml_template.txt", "r")
    as_string = ""
    for line in file_in:
        as_string += (line + "\n")

    return as_string

def colour_list_generator(species_list):
    colour_list = {}
    for species in species_list:
        colour_list[species] = colour_gen(species)  ## aabbrrgg

    return colour_list

def colour_gen(word):
    int_from_str = 0
    for i in range(len(word)):
        int_from_str += ord(word[i])
    colour = "#ff"
    for i in range(3):
        value = (int_from_str >> (i*2)) & 0xFF    ## Can mess about with the
        to_add = hex(value).strip("0x")           ## bitshifting here to change
        if len(to_add) == 1:                      ## colours (change what i is
            to_add += "0"                         ## multiplied by)
        colour+=(to_add)
    return colour


def auto_column(column_headers):
    auto_columns = {}

    for column in column_headers:
        if column[1].upper() == "SITE":
            auto_columns["site_col"] = column[0]
        elif column[1].upper() == "TRANSECT":
            auto_columns["tran_col"] = column[0]
        elif column[1].upper() == "REC_TIME":
            auto_columns["time_col"] = column[0]
        elif column[1].upper() == "LAT":
            auto_columns["lat_col"] = column[0]
        elif column[1].upper() == "LONG":
            auto_columns["long_col"] = column[0]
        elif column[1].upper() == "DATE":
            auto_columns["date_col"] = column[0]

    return(auto_columns)


def menu():
    print("1: Load file")
    print("2: Pick sheet")
    print("3: Set columns")
    print("4: Check data")
    print("5: Make maps")
    print("6: Colour generator")
    print("0: Exit")

    user_input = int(input("What would you like to do? "))

    return user_input

def main():

    choice = 1

    while(choice != 0):
        choice = menu()
        if choice == 1:
            current_wb = get_sheet()
        if choice == 2:
            try:
                sheet_list = display_sheets(current_wb)
            except(NameError):
                print("Please load a sheet")
                current_wb = get_sheet()
                sheet_list = display_sheets(current_wb)
            sheet_num = int(input("Which sheet would you like to load? "))
            current_sheet = current_wb[sheet_list[sheet_num-1]]
        if choice == 3:
            col_list = show_columns(current_sheet)
            auto_cols = auto_column(col_list)
            for column in auto_cols:
                print(("{} : {}").format(auto_cols[column]+1,column))
            correct = False
            while correct == False:
                answer = input("Are these columns correct? Y/N ")
                if answer[0].upper() == "Y":
                    correct = True
                elif answer[0].upper() == "N":
                    correct = False
                    for column in auto_cols:
                        new_col = int(input("Reassigning {} from {} to: ".format(column, auto_cols[column]+1)))
                        auto_cols[column] = new_col-1
                        print(("{} : {}").format(auto_cols[column]+1,column))
                else:
                    print("Input must be Y or N.")

            ### Assigning species columns

            species_cols = []
            num_spec_cols = int(input("How many species columns are there?"))
            for i in range(num_spec_cols):
                species_col = int(input("Enter species column {0}:".format(i+1)))
                species_cols.append(species_col)

        if choice == 4:
            columns_str = input("Please enter the columns whose items you wish to see (comma-separated):")

            columns = columns_str.split(",")

            for column in columns:
                column_items = get_column_items(current_sheet, int(column) -1 )
                print(column_items)

        if choice == 5:
            mk_kml(current_sheet,
                   species_cols,
                   auto_cols["site_col"],
                   auto_cols["time_col"],
                   auto_cols["tran_col"],
                   10,
                   auto_cols["lat_col"],
                   auto_cols["long_col"],
                   colour_list
                   )

        if choice == 6:
            species_list = get_column_items(current_sheet, species_cols[0]-1)
            print(species_list)
            colour_list = colour_list_generator(species_list)
            print(colour_list)





    '''


    sheet_num = int(input("Please enter the sheet number"))

    sheet = load_sheet(wb, sheet_list[sheet_num-1])

    column_headers = show_columns(sheet)

    columns_str = input("Please enter the columns whose items you wish to see (comma-separated):")

    columns = columns_str.split(",")

    for column in columns:
        column_items = get_column_items(sheet, int(column) -1 )
        print(column_items)

    species_cols = []
    num_spec_cols = int(input("How many species columns are there?"))
    for i in range(num_spec_cols):
        species_col = int(input("Enter species column {0}:".format(i+1)))
        species_cols.append(species_col)

    google_earth_colours = ["grn","red","ylw","blue","purple","pink","ltblu","wht"] ## these are hardcoded by google. I think a more elegant solution exists.

    species_list = get_column_items(sheet, species_cols[0])
    species_colour = {}

    ### WARNING! AT PRESENT THIS ONLY HAS CAPACITY FOR 8 DIFFERENT SPECIES!!!
    for i in range(len(species_list)):
        if i < 8:
            j = i
        else:
            j=7
        species_colour[species_list[i]] = google_earth_colours[j]

    print(auto_column(show_columns(sheet)))

    mk_kml(sheet,
           species_cols,
           )
    '''







main()




